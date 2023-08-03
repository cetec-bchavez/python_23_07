from openpyxl import load_workbook

wb = load_workbook('facturacion.xlsx')

sheet1 = wb.active

rowx=5

for rowx in range(2,rowx):
    sheet1.cell(row=rowx,column=3).value=sheet1.cell(row=rowx,column=3).value+100
    rowx+=1

wb.save('facturacion2.xlsx')