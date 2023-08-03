from openpyxl import Workbook
from base.factura import Factura

wb = Workbook()

sheet1 = wb.active

sheet1['A1'] = 'Id'
sheet1['B1'] = 'Producto'
sheet1['C1'] = 'Cantidad'
sheet1['D1'] = 'Precio'
sheet1['E1'] = 'Total'

facturas = list()

factura1:Factura = None

print('------------------------- 1 OBJETO ------------------------------')
# OBJETOS
# nombre = Clase(valor1,valor2,valor3)

factura1 = Factura(0,'',0,0.0,0.0)

# Valores a Parametros
factura1.producto = 'Mouse'
factura1.cantidad = 5
factura1.precio = 7

# Procesar Calculo
factura1.calcularTotal()

print(factura1.producto,factura1.cantidad,factura1.total)


print('------------------------- LISTA OBJETOS ------------------------------')
factura2 = Factura(0,'Teclado',7,8.0,0.0)
factura2.calcularTotal()

factura3 = Factura(0,'Monitor',9,3.0,0.0)
factura3.calcularTotal()


facturas = list()
#facturas = []

facturas.append(factura1)
facturas.append(factura2)
facturas.append(factura3)

print(facturas)

rowx=2
facturax:Factura = None

for facturax in facturas:

    sheet1.cell(row=rowx,column=1).value=facturax.id
    sheet1.cell(row=rowx,column=2).value=facturax.producto
    sheet1.cell(row=rowx,column=3).value=facturax.cantidad
    sheet1.cell(row=rowx,column=4).value=facturax.precio
    sheet1.cell(row=rowx,column=5).value=facturax.total

    rowx+=1

wb.save('facturacion.xlsx')